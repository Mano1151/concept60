from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

folder = 'Vulnerability Test Results'
filename = os.path.join(folder, 'Backend_Security_Review.xlsx')
summary_filename = os.path.join(folder, 'Executive_Concise_Summary.txt')

findings = [
    ('Critical', 'server/.env', 'Sensitive Data Exposure', 'Hardcoded secrets and private keys are present in backend environment file.', 'Remove secrets from source tree, use secret manager or deploy-time env vars, and ensure .env is not committed.'),
    ('High', 'server/services/claudeService.js', 'Prompt Injection / AI Injection', 'User-controlled prompt data is embedded into LLM prompts with weak pattern filtering, enabling model-level injection and unsafe output.', 'Apply stronger prompt sanitization, strict output validation, and avoid relying on model instruction filtering alone; use allowlists and external prompt wrappers.'),
    ('Medium', 'server/services/claudeService.js', 'Sensitive Data Exposure in Logs', 'Raw AI responses and final keywords are logged to console, risking leakage of sensitive or proprietary data.', 'Stop logging raw model responses and user-generated content in production logs; sanitize or remove debug logs.'),
    ('Medium', 'server/index.js', 'Authentication / Session Management', 'No logout or token revocation endpoint exists and long-lived Firebase tokens may remain valid if compromised.', 'Implement explicit token revocation support or require shorter-lived tokens, and document logout flow for clients.'),
    ('Medium', 'server/index.js', 'API Security', 'No dedicated rate limiter is configured for /api/history route; only global limit applies.', 'Add route-specific rate limiting for history and other sensitive endpoints to harden abuse resistance.'),
    ('Low', 'server/index.js', 'Infrastructure Configuration', 'trust proxy is set globally without explicit proxy validation, which can misidentify client IPs if not behind a trusted proxy.', 'Verify proxy environment and use precise trust proxy settings to prevent IP spoofing for rate limiting and logging.'),
    ('Low', 'server/index.js', 'API Security', 'CORS is restricted but does not include logging of blocked origins or monitoring for misconfiguration.', 'Keep origin policy strict and monitor CORS failures; confirm allowedOrigin is not injected from unsafe sources.'),
    ('Low', 'server/routes/history.js', 'Authorization / Access Control', 'History deletion is limited to authenticated users only, but there is no audit of user permissions or user role assertions.', 'Maintain the pattern, but add RBAC or ownership assertion checks if user roles or admin actions are introduced.'),
]

wb = Workbook()
ws = wb.active
ws.title = 'Findings'
headers = ['Severity', 'File Path', 'Vulnerability Type', 'Brief Explanation', 'Recommended Remediation']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row, finding in enumerate(findings, start=2):
    for col, value in enumerate(finding, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for column_cells in ws.columns:
    length = max(len(str(cell.value or '')) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 15), 60)

ws2 = wb.create_sheet('EndPoints')
ws2.append(['Endpoint', 'Authentication Required', 'Notes'])
rows = [
    ('GET /', 'No', 'Public health check only.'),
    ('POST /api/concept', 'Yes', 'Requires Firebase ID token. Good authentication enforcement.'),
    ('POST /api/video', 'Yes', 'Requires Firebase ID token. Good authentication enforcement.'),
    ('POST /api/qa/pdf-question', 'Yes', 'Requires Firebase ID token. Good authentication enforcement.'),
    ('GET /api/history', 'Yes', 'Requires Firebase ID token. No route-specific rate limiter.'),
    ('DELETE /api/history/:entryId', 'Yes', 'Requires Firebase ID token. Uses user-scoped Firestore path.'),
]
for row_data in rows:
    ws2.append(row_data)
for cell in ws2[1]:
    cell.font = Font(bold=True)
wb.save(filename)

summary = '''Executive Concise Summary

Critical Findings:
- Hardcoded secrets in backend .env file expose API keys and Firebase service account private key. Remove all secrets from source files and use secure runtime secret management.
- User input is directly embedded into LLM prompts in server/services/claudeService.js. Current prompt injection checks are insufficient for model-level injection attacks.

High/Medium Findings:
- Console logs expose raw AI output and generated keyword data.
- No logout/token revocation endpoint exists; stolen Firebase tokens remain valid until expiry.
- /api/history lacks dedicated route-specific rate limiting, creating an abuse path.
- trust proxy is globally enabled without explicit proxy validation.

Recommendations:
- Remove .env from repository and use secret storage.
- Harden LLM prompt handling and validate model output strictly.
- Remove sensitive logs from production.
- Add explicit logout/revocation support and minimize token lifetime.
- Add route-specific rate limiting for sensitive endpoints.
'''
with open(summary_filename, 'w') as f:
    f.write(summary)
print('Generated:', filename)
