from openpyxl import load_workbook

src = r'C:\Users\Mano\OneDrive\Desktop\concept60\Vulnerability Test Results\test-case-value-model-copy.xlsx'
dst = r'C:\Users\Mano\OneDrive\Desktop\concept60\Vulnerability Test Results\test-case-value-model-updated.xlsx'

wb = load_workbook(src)
ws = wb['TestCases']
rows = list(ws.iter_rows(values_only=True))
headers = list(rows[0])
if 'Status' not in headers:
    headers.insert(9, 'Status')

new_rows = [tuple(headers)]
for row in rows[1:]:
    row_list = list(row)
    if row_list[0] == 'TC007':
        row_list[2] = '/api/history/:entryId'
        row_list[3] = 'DELETE'
        row_list[4] = 'Yes'
        if len(row_list) < len(headers):
            row_list.insert(9, 'Open')
    else:
        if len(row_list) < len(headers):
            row_list.insert(9, 'Open')
    new_rows.append(tuple(row_list))

for r_idx, row in enumerate(new_rows, start=1):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val)

# Set column widths generically
for c_idx in range(1, len(headers) + 1):
    col_letter = ws.cell(row=1, column=c_idx).column_letter
    ws.column_dimensions[col_letter].width = 20

wb.save(dst)
print('Saved updated workbook to', dst)
