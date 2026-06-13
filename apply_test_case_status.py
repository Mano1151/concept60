from openpyxl import load_workbook

src = r'C:\Users\Mano\OneDrive\Desktop\concept60\Vulnerability Test Results\test-case-value-model-updated.xlsx'
dst = r'C:\Users\Mano\OneDrive\Desktop\concept60\Vulnerability Test Results\test-case-value-model-fixed.xlsx'

wb = load_workbook(src)
ws = wb['TestCases']
for row in ws.iter_rows(min_row=2, max_col=11, values_only=False):
    if row[0].value == 'TC007':
        ws.cell(row=row[0].row, column=10, value='Fixed')
    elif row[0].value is not None:
        if ws.cell(row=row[0].row, column=10).value is None:
            ws.cell(row=row[0].row, column=10, value='Open')

wb.save(dst)
print('Saved fixed workbook to', dst)
