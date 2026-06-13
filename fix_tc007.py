from openpyxl import load_workbook

path = r'C:\Users\Mano\OneDrive\Desktop\concept60\Vulnerability Test Results\test-case-value-model-updated.xlsx'
wb = load_workbook(path)
ws = wb['TestCases']
for row in ws.iter_rows(min_row=2, values_only=False):
    if row[0].value == 'TC007':
        corrected = [
            'TC007',
            'Authorization',
            '/api/history/:entryId',
            'DELETE',
            'Yes',
            'entryId not owned by user',
            'user cannot delete other user history',
            'Access denied for other user history',
            'High',
            'Open',
            'Ensure only owner can delete',
        ]
        for i, val in enumerate(corrected, start=1):
            ws.cell(row=row[0].row, column=i, value=val)
        break
wb.save(path)
print('Corrected TC007 row in', path)
