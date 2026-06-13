import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

workbook = Workbook()
value_sheet = workbook.active
value_sheet.title = 'ValueModel'
test_sheet = workbook.create_sheet('TestCases')

value_headers = ['Field Name', 'Data Type', 'Valid Value', 'Invalid Value', 'Edge Case', 'Test Purpose']
value_rows = [
    ['concept', 'string', 'Gravity', '', 'Whitespace-only', 'Validate prompt field length and content'],
    ['category', 'string', 'Physics', '', 'Special chars / <script>', 'Validate category constraints'],
    ['pdfText', 'string', 'Long document text', '', 'Very large text > 30k chars', 'Validate PDF source sanitization and length limits'],
    ['question', 'string', 'What is force?', '', 'Prompt injection payload', 'Validate question input and model safety'],
    ['entryId', 'string', 'abcd1234', '', 'Missing or empty path param', 'Validate history delete authorization'],
    ['Authorization', 'header', 'Bearer valid.token', 'Missing token', 'Malformed or expired token', 'Verify auth enforcement'],
]

value_sheet.append(value_headers)
for row in value_rows:
    value_sheet.append(row)

for i, column in enumerate(value_headers, 1):
    value_sheet.column_dimensions[get_column_letter(i)].width = 20

case_headers = ['Test Case ID', 'Area', 'Endpoint', 'HTTP Method', 'Auth Required', 'Input', 'Input Value / Scenario', 'Expected Result', 'Severity', 'Notes']
case_rows = [
    ['TC001', 'Authentication', '/api/concept', 'POST', 'Yes', 'Missing Authorization header', 'no Authorization header', '401 Unauthorized', 'High', 'Ensure public AI generation endpoint requires valid token'],
    ['TC002', 'Validation', '/api/concept', 'POST', 'Yes', 'concept too short', 'concept="A"', '400 Bad Request', 'Medium', 'Reject short concept values'],
    ['TC003', 'Validation', '/api/video', 'POST', 'Yes', 'category invalid type', 'category=123', '400 Bad Request', 'Medium', 'Reject non-string category values'],
    ['TC004', 'Authentication', '/api/qa/pdf-question', 'POST', 'Yes', 'valid auth, missing question', 'pdfText provided, question missing', '400 Bad Request', 'High', 'Require both fields'],
    ['TC005', 'Injection', '/api/qa/pdf-question', 'POST', 'Yes', 'prompt injection payload in question', 'question="<script>alert(1)</script>"', 'Sanitized and processed without injection', 'High', 'Ensure model prompt sanitization is applied'],
    ['TC006', 'Rate Limiting', '/api/concept', 'POST', 'Yes', 'repeated requests', '21 requests within 1 minute', '429 Too Many Requests', 'Medium', 'Verify endpoint-specific quota'],
    ['TC007', 'Authorization', '/api/history', 'DELETE /:entryId', 'DELETE', 'Yes', 'entryId not owned by user', 'user cannot delete other user history', 'High', 'Ensure only owner can delete'],
    ['TC008', 'Input Size', '/api/qa/pdf-question', 'POST', 'Yes', 'pdfText too large', 'pdfText length 30001', '400 Bad Request', 'Medium', 'Reject oversized payloads'],
    ['TC009', 'Model Safety', '/api/concept', 'POST', 'Yes', 'concept with control characters', 'concept contains null bytes and <>', 'Sanitized prompt input', 'Medium', 'Verify prompt sanitization'],
    ['TC010', 'Logging', '/api/concept', 'POST', 'Yes', 'successful request', 'valid payload', 'No raw AI response logged', 'Low', 'Ensure sensitive output is not logged in production'],
]

test_sheet.append(case_headers)
for row in case_rows:
    test_sheet.append(row)

for i, column in enumerate(case_headers, 1):
    test_sheet.column_dimensions[get_column_letter(i)].width = 22

workbook.save(r'Vulnerability Test Results/test-case-value-model.xlsx')
print('Created Vulnerability Test Results/test-case-value-model.xlsx')
