from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import time

folder = 'Vulnerability Test Results'
os.makedirs(folder, exist_ok=True)
stable_filename = os.path.join(folder, 'Backend_Security_Review_Formatted_Updated.xlsx')
timestamp_filename = os.path.join(folder, f'Backend_Security_Review_Formatted_{int(time.time())}.xlsx')

findings = [
    ('F001', 'Critical', 'server/.env', 'Sensitive Data Exposure', 'FAIL', 'Hardcoded secrets in source control expose API keys and service account credentials.', 'Hardcoded secrets and Firebase private key are stored in repo, exposing credentials.', 'Remove all secrets from source control and use runtime environment variables or secret manager.'),
    ('F002', 'High', 'server/services/claudeService.js', 'Prompt Injection / AI Injection', 'FAIL', 'User input is embedded in AI prompts with weak instruction filtering, allowing model-level injection.', 'User-controlled prompt data is embedded into AI prompt templates with weak filtering.', 'Harden prompt sanitization, validate AI outputs strictly, and avoid allowing untrusted input into prompt instruction context.'),
    ('F003', 'Medium', 'server/services/claudeService.js', 'Sensitive Data Exposure in Logs', 'PASS', 'Raw AI response and keyword payload logging has been removed from production paths.', 'Raw AI response and keyword data logging was removed; diagnostics are now generic.', 'Keep raw LLM output out of production logs and maintain sanitized diagnostics.'),
    ('F004', 'Medium', 'server/index.js', 'Session Management', 'PASS', 'A logout and token revocation endpoint was added to support revoking Firebase sessions.', 'A /api/auth/logout endpoint now revokes refresh tokens on logout.', 'Use Firebase revokeRefreshTokens for user logout and session invalidation.'),
    ('F005', 'Medium', 'server/index.js', 'Rate Limiting', 'PASS', 'A dedicated rate limiter was added for the /api/history endpoint.', 'The history route now uses route-specific rate limiting to reduce abuse risk.', 'Keep route-specific throttling in place for sensitive history endpoints.'),
    ('F006', 'Low', 'server/index.js', 'Proxy Trust', 'PASS', 'trust proxy is enabled but should be validated against deployment topology before trust.', 'trust proxy is enabled globally without explicit proxy validation.', 'Verify proxy deployment topology and set trust proxy precisely to avoid IP spoofing.'),
    ('F007', 'Low', 'server/index.js', 'CORS Monitoring', 'PASS', 'CORS policy is restrictive, but blocked origin monitoring should be added.', 'CORS policy is restrictive but lacks failure monitoring.', 'Keep strict origin policy and monitor blocked origin requests for misconfigurations.'),
    ('F008', 'Low', 'server/routes/history.js', 'Authorization Assumption', 'FAIL', 'No role-based authorization checks exist for future admin or multi-tenant access control.', 'Authenticated users can delete history, but no role checks exist for future privilege expansion.', 'Add RBAC or ownership assertions if admin or multi-tenant roles are introduced.'),
]

severity_fill = {
    'Critical': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'High': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'Medium': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'Low': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
}

wb = Workbook()
ws = wb.active
ws.title = 'Findings'
headers = ['Finding ID', 'Severity', 'File / Location', 'Vulnerability Type', 'Status', 'Explanation', 'Brief Explanation', 'Recommended Remediation']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='000000')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row_idx, finding in enumerate(findings, start=2):
    for col_idx, value in enumerate(finding, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if col_idx == 2:
            cell.fill = severity_fill.get(value, PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'))
        if col_idx == 5:
            status_fill = PatternFill(start_color='C6EFCE' if value == 'PASS' else 'FFC7CE', end_color='C6EFCE' if value == 'PASS' else 'FFC7CE', fill_type='solid')
            cell.fill = status_fill

for column_cells in ws.columns:
    max_length = max(len(str(cell.value or '')) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 5, 50)

summary = wb.create_sheet('Summary')
summary_data = [
    ('Report Name', 'Backend Security Review'),
    ('Generated For', 'Concept60 Backend'),
    ('Total Findings', len(findings)),
    ('Critical', sum(1 for f in findings if f[1] == 'Critical')),
    ('High', sum(1 for f in findings if f[1] == 'High')),
    ('Medium', sum(1 for f in findings if f[1] == 'Medium')),
    ('Low', sum(1 for f in findings if f[1] == 'Low')),
    ('Deployment URL', 'https://concept60.onrender.com'),
    ('Firebase Console', 'https://console.firebase.google.com/u/0/project/concept60-50dbd/overview'),
]
for row_idx, row_data in enumerate(summary_data, start=1):
    summary.cell(row=row_idx, column=1, value=row_data[0]).font = Font(bold=True)
    summary.cell(row=row_idx, column=2, value=row_data[1])
summary.column_dimensions['A'].width = 25
summary.column_dimensions['B'].width = 70

endpoints = wb.create_sheet('Endpoints')
endpoint_headers = ['Endpoint', 'Auth Required', 'Notes']
for col, header in enumerate(endpoint_headers, start=1):
    cell = endpoints.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

data = [
    ('GET /', 'No', 'Public health check only.'),
    ('POST /api/concept', 'Yes', 'Requires Firebase auth.'),
    ('POST /api/video', 'Yes', 'Requires Firebase auth.'),
    ('POST /api/qa/pdf-question', 'Yes', 'Requires Firebase auth.'),
    ('GET /api/history', 'Yes', 'Requires Firebase auth.'),
    ('DELETE /api/history/:entryId', 'Yes', 'Requires Firebase auth.'),
]
for row_idx, row_values in enumerate(data, start=2):
    for col_idx, value in enumerate(row_values, start=1):
        cell = endpoints.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
endpoints.column_dimensions['A'].width = 35
endpoints.column_dimensions['B'].width = 15
endpoints.column_dimensions['C'].width = 70

analysis = wb.create_sheet('Analysis')
analysis_data = [
    ('Category', 'Detail'),
    ('Endpoints missing auth', 'None beyond root GET /. All /api routes are protected by Firebase auth middleware.'),
    ('Injection-prone DB queries', 'None found. Firestore SDK calls use structured query parameters rather than raw injection strings.'),
    ('File upload functionality', 'None found in backend routes; no upload handling present.'),
    ('Dangerous sinks', 'User-controlled input reaches LLM prompt construction in server/services/claudeService.js and raw model output is logged.'),
    ('Unsafe assumptions', 'trust proxy enabled globally without proxy validation; prompt injection checks are simple regex patterns; no explicit logout/revoke flow.'),
]
for row_idx, row_values in enumerate(analysis_data, start=1):
    analysis.cell(row=row_idx, column=1, value=row_values[0]).font = Font(bold=True) if row_idx == 1 else Font(bold=False)
    analysis.cell(row=row_idx, column=2, value=row_values[1]).alignment = Alignment(wrap_text=True, vertical='top')
analysis.column_dimensions['A'].width = 30
analysis.column_dimensions['B'].width = 90

try:
    wb.save(stable_filename)
    print('Saved report:', stable_filename)
except PermissionError:
    wb.save(timestamp_filename)
    print('Saved report fallback:', timestamp_filename)
