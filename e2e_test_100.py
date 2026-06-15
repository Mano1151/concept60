"""
Selenium E2E - 100 Test Cases Runner for Concept60

This script programmatically runs 100 lightweight end-to-end checks
against the deployed app and writes results to an Excel workbook.
"""

import time
import logging
from datetime import datetime
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Basic logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class BulkE2ETests:
    def __init__(self, base_url="https://concept60.onrender.com"):
        self.base_url = base_url
        self.driver = None
        self.test_results = []
        self.test_count = 0
        self.passed_count = 0
        self.failed_count = 0

    def setup_driver(self):
        try:
            chrome_options = Options()
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            # Run visible by default to help debugging; comment out to run headless
            # chrome_options.add_argument('--headless=new')

            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.set_page_load_timeout(30)
            return True
        except Exception as e:
            logger.error(f"Failed to initialize WebDriver: {e}")
            return False

    def log_test(self, name, status, description=""):
        self.test_count += 1
        if status.upper() == 'PASS':
            self.passed_count += 1
        else:
            self.failed_count += 1

        self.test_results.append({
            'Test ID': f'TC_{self.test_count:03d}',
            'Test Name': name,
            'Status': status.upper(),
            'Description': description,
            'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        logger.info(f"{name} - {status} - {description}")

    def single_iteration_check(self, index):
        """Per-iteration composite check that bundles a few lightweight verifications."""
        name = f"Iteration {index} composite checks"
        failures = []

        try:
            # Load base page
            self.driver.get(self.base_url)
            wait = WebDriverWait(self.driver, 10)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

            # Title check
            title = (self.driver.title or '').strip()
            if not title or title.lower() == 'about:blank':
                failures.append('Missing or invalid title')

            # Navbar check
            nav_found = len(self.driver.find_elements(By.TAG_NAME, 'nav')) > 0 or len(self.driver.find_elements(By.CLASS_NAME, 'navbar')) > 0
            if not nav_found:
                failures.append('Navbar missing')

            # Search interaction: send a query based on index
            search_inputs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='search'], input[placeholder*='search' i], input[type='text']")
            if search_inputs:
                try:
                    q = f"automation test {index}"
                    el = search_inputs[0]
                    el.clear()
                    el.send_keys(q)
                    time.sleep(0.5)
                    val = (el.get_attribute('value') or '').lower()
                    if q.split()[-1] not in val:
                        failures.append('Search input did not accept text')
                except Exception:
                    failures.append('Search input interaction failed')
            else:
                failures.append('No search input')

            # Link click attempt: try to click a link at pseudo-random index
            links = self.driver.find_elements(By.TAG_NAME, 'a')
            if links:
                idx = index % max(1, len(links))
                try:
                    link = links[idx]
                    if link.is_displayed() and link.is_enabled():
                        # safe click: open in same tab
                        try:
                            link.click()
                            time.sleep(0.5)
                            self.driver.back()
                        except Exception:
                            # fallback: navigate to href
                            href = link.get_attribute('href')
                            if href:
                                self.driver.get(href)
                                time.sleep(0.5)
                                self.driver.back()
                except Exception:
                    failures.append('Link interaction failed')

            # Responsive check: toggle window size every other test
            if index % 2 == 0:
                self.driver.set_window_size(375, 667)
            else:
                self.driver.set_window_size(1366, 768)
            time.sleep(0.2)

            # Basic accessibility checks
            imgs = self.driver.find_elements(By.TAG_NAME, 'img')
            imgs_with_alt = len([i for i in imgs if i.get_attribute('alt')])
            if imgs and imgs_with_alt == 0:
                failures.append('Images missing alt text')

        except Exception as e:
            failures.append(f'Exception during iteration: {e}')

        if failures:
            self.log_test(name, 'FAIL', '; '.join(failures))
            return False
        else:
            self.log_test(name, 'PASS', 'All composite checks passed')
            return True

    def run_100(self):
        logger.info('Starting 100 E2E iterations')
        if not self.setup_driver():
            logger.error('WebDriver setup failed; aborting')
            return False

        try:
            for i in range(1, 101):
                try:
                    self.single_iteration_check(i)
                    time.sleep(0.25)  # small pause between iterations
                except Exception as e:
                    logger.error(f'Iteration {i} crashed: {e}')
            # After all runs, generate report
            report_path = self.generate_excel_report()
            logger.info(f'Report written: {report_path}')
            logger.info(f'Total: {self.test_count}, Passed: {self.passed_count}, Failed: {self.failed_count}')
            return True
        finally:
            try:
                self.driver.quit()
            except Exception:
                pass

    def generate_excel_report(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Results'

        # Styles
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        pass_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        fail_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        thin_border = Border()

        headers = ['Test ID', 'Test Name', 'Status', 'Description', 'Timestamp']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        for r, res in enumerate(self.test_results, start=2):
            ws.cell(row=r, column=1, value=res['Test ID'])
            ws.cell(row=r, column=2, value=res['Test Name'])
            ws.cell(row=r, column=3, value=res['Status'])
            ws.cell(row=r, column=4, value=res['Description'])
            ws.cell(row=r, column=5, value=res['Timestamp'])

            status_cell = ws.cell(row=r, column=3)
            if res['Status'] == 'PASS':
                status_cell.fill = pass_fill
            else:
                status_cell.fill = fail_fill

        # Summary sheet
        s = wb.create_sheet('Summary')
        s['A1'] = 'Total Tests'
        s['B1'] = self.test_count
        s['A2'] = 'Passed'
        s['B2'] = self.passed_count
        s['A3'] = 'Failed'
        s['B3'] = self.failed_count
        s['A4'] = 'Execution Date'
        s['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        s['A5'] = 'URL'
        s['B5'] = self.base_url

        report_name = f"Test_Report_100_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = Path.cwd() / report_name
        wb.save(report_path)
        return report_path


def main():
    tester = BulkE2ETests()
    tester.run_100()


if __name__ == '__main__':
    main()
